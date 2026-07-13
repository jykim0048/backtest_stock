#!/usr/bin/env python3
"""KRX 전종목 업종분류 엑셀 → KOSPI 섹터→대표 구성종목 매핑 JSON 생성.

KRX 데이터포털 [업종분류 현황] 다운로드 엑셀(컬럼: 종목코드·종목명·시장구분·업종명·
종가·대비·등락률·시가총액)을 입력받아, 장중 시황의 KIS 업종 섹터 히트에 '관련주'를
붙이기 위한 매핑을 만든다. KIS 업종 지수(FHPUP02140000)와 KRX 업종분류는 같은 KRX
산업 분류라 업종명이 대체로 일치하지만, 표기 차(공백·중점·'기/기기')가 있어 정규화 키로
저장한다(파이프라인이 KIS 업종명도 같은 규칙으로 정규화해 조인).

사용:
  python build_krx_sector_map.py <KOSPI_xlsx> [--out public/assets/krx_sector_map.json]

KRX 분류는 자주 바뀌지 않으므로 산출 JSON 만 커밋하고, 갱신 시 재다운로드→재실행한다.
"""
import os
import re
import sys
import json
import argparse

import openpyxl

ROOT = os.path.dirname(os.path.abspath(__file__))
TOP_PER_SECTOR = 12   # 섹터당 시총 상위 N개 보관(파이프라인이 4개 사용, 여유분 확보)


def norm_sector(name):
    """업종명 정규화 키 — 공백·중점(·・)·괄호 제거. KIS/KRX 표기차 흡수용.
    '의료·정밀기기'/'의료·정밀기' 처럼 '기기/기' 꼬리 차이는 파이프라인의 접두 매칭이 흡수."""
    return re.sub(r"[\s·・()]", "", str(name or "")).strip()


# 시장구분 → 저장 필드. KOSPI 는 표시·지수 정합용(stocks), KOSDAQ 은 테마 매칭
# 확장 전용(kosdaqStocks) — 시총 스케일이 달라 통합 정렬하면 컷(top12)이 왜곡되므로
# 시장별로 각각 top12 를 보관한다(2026-07-13, 기계·장비/제약 케이스).
_MARKET_FIELD = {"KOSPI": "stocks", "KOSDAQ": "kosdaqStocks"}

# KIS 산업별 지수의 '대분류'(제조·금융)는 KRX 세부 업종분류에 대응 키가 없어 관련주가
# 비던 문제(2026-07-13) — 세부 업종 합집합에서 시총 상위로 합성 엔트리를 만든다.
# 키는 정규화 업종명(norm_sector). 멤버에 없는 키는 무시(시장별 분류 차 흡수).
_COMPOSITE = {
    "금융": {"name": "금융",
             "members": ["은행", "증권", "보험", "기타금융", "금융"]},
    "제조": {"name": "제조",
             "members": ["음식료·담배", "섬유·의류", "종이·목재", "화학", "제약", "비금속",
                          "금속", "기계·장비", "전기·전자", "의료·정밀기기", "운송장비·부품",
                          "기타제조", "제조"]},
}


def load_index_composition(path):
    """KIS 산업별 '지수 구성종목' 엑셀(종목코드*001·업종명·시총(상장)·비중, 비중순) 파싱.
    업종분류 엑셀과 다른 실제 지수 정합 소스 — 대분류(금융·제조) 지수는 업종분류
    체계와 구성 기준이 달라(2026-07-13 실측: 금융 지수 = 금융98+증권29+보험14)
    이 파일이 정답이다. 반환: [{code,name,cap}] (우선주 제외, 시총 내림차순)."""
    wb = openpyxl.load_workbook(path, data_only=True)
    out = []
    for row in wb.active.iter_rows(min_row=2, values_only=True):
        raw_code, name, cap = row[0], row[2], row[6]
        if not raw_code or not name:
            continue
        code = str(raw_code).split("*")[0].strip().zfill(6)
        if not code.isdigit() or not code.endswith("0"):   # 우선주(끝자리≠0) 제외
            continue
        try:
            c = float(str(cap).replace(",", "")) if cap is not None else 0.0
        except (ValueError, TypeError):
            c = 0.0
        out.append({"code": code, "name": str(name).strip(), "cap": c})
    out.sort(key=lambda s: s["cap"], reverse=True)
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", nargs="*", default=[],
                    help="KRX 업종분류 엑셀 경로(코스피/코스닥, 여러 파일 가능)")
    ap.add_argument("--index-xlsx", action="append", default=[], metavar="업종명=경로",
                    help="KIS 지수 구성종목 엑셀로 해당 섹터의 KOSPI 관련주를 직접 지정 "
                         "(예: --index-xlsx 금융=금융업종구성.xlsx) — 합성 추정보다 우선")
    ap.add_argument("--out", default=os.path.join(ROOT, "public", "assets", "krx_sector_map.json"))
    ap.add_argument("--date", default="")
    args = ap.parse_args()

    # 입력 엑셀에 없는 시장은 기존 맵에서 이월한다 — 예: 코스닥 엑셀만 새로 받아
    # kosdaqStocks 만 갱신하고 기존 KOSPI stocks 는 유지.
    by_sector, covered = {}, set()
    try:
        with open(args.out, encoding="utf-8") as f:
            for k, e in ((json.load(f) or {}).get("sectors") or {}).items():
                by_sector[k] = {fld: list(v) if isinstance(v, list) else v
                                for fld, v in e.items()}
    except Exception:
        pass

    raw = {}   # (key, field) → [{code,name,cap}]
    for path in args.xlsx:
        wb = openpyxl.load_workbook(path, data_only=True)
        for row in wb.active.iter_rows(min_row=2, values_only=True):
            code, name, market, upjong = row[0], row[1], row[2], row[3]
            field = _MARKET_FIELD.get(str(market or "").strip().upper())
            if not field or not code or not upjong:
                continue
            try:
                cap = float(str(row[7]).replace(",", "")) if row[7] is not None else 0.0
            except (ValueError, TypeError):
                cap = 0.0
            key = norm_sector(upjong)
            covered.add(field)
            by_sector.setdefault(key, {"name": str(upjong).strip()})
            raw.setdefault((key, field), []).append(
                {"code": str(code).strip().zfill(6), "name": str(name).strip(), "cap": cap})

    # 대분류 합성(금융·제조) — 이번 입력 엑셀이 커버하는 시장(field)에 한해, 멤버 세부
    # 업종의 raw 행(시총 보유)을 합쳐 시총 상위로 재산출. 입력에 없는 시장은 기존 이월.
    for ckey, spec in _COMPOSITE.items():
        member_keys = {norm_sector(m) for m in spec["members"]}
        for field in set(f for (_, f) in raw.keys()):
            pool = [s for (k, f), stocks in raw.items()
                    if f == field and k in member_keys and k != ckey for s in stocks]
            # 코스닥 엑셀처럼 '금융' 자체가 세부 업종으로 오는 경우는 그대로 포함
            pool += list(raw.get((ckey, field), []))
            if pool:
                seen_c, dedup = set(), []
                for s in pool:
                    if s["code"] not in seen_c:
                        seen_c.add(s["code"])
                        dedup.append(s)
                raw[(ckey, field)] = dedup
                by_sector.setdefault(ckey, {"name": spec["name"]})

    for (key, field), stocks in raw.items():
        stocks.sort(key=lambda s: s["cap"], reverse=True)
        by_sector[key][field] = [{"code": s["code"], "name": s["name"]}
                                 for s in stocks[:TOP_PER_SECTOR]]

    # 지수 구성종목 엑셀 — 해당 섹터의 KOSPI 관련주를 지수 정합 그대로 채운다(최우선).
    for spec in args.index_xlsx:
        name, _, path = spec.partition("=")
        name = name.strip()
        rows = load_index_composition(path.strip())
        if not rows:
            print(f"  [index-xlsx] {name}: 구성종목 파싱 실패 — 건너뜀")
            continue
        key = norm_sector(name)
        by_sector.setdefault(key, {"name": name})
        by_sector[key]["stocks"] = [{"code": s["code"], "name": s["name"]}
                                    for s in rows[:TOP_PER_SECTOR]]
        raw.pop((key, "stocks"), None)      # 합성 추정치가 있으면 지수 정합본으로 대체
        print(f"  [index-xlsx] {name}: {len(rows)}종목(우선주 제외) → KOSPI top{TOP_PER_SECTOR}")

    date = args.date
    if not date:
        src0 = (args.xlsx or [p.partition('=')[2] for p in args.index_xlsx] or ["" ])[0]
        m = re.search(r"(\d{8})", os.path.basename(src0))
        date = f"{m.group(1)[:4]}-{m.group(1)[4:6]}-{m.group(1)[6:]}" if m else ""

    out = {"date": date,
           "note": "KRX 업종분류 — KIS 업종 섹터 히트의 관련주 매핑. 정규화 업종명 키. "
                   "stocks=KOSPI 시총상위(표시·지수 정합), kosdaqStocks=KOSDAQ 시총상위(테마 매칭 확장).",
           "sectors": by_sector}
    os.makedirs(os.path.dirname(args.out), exist_ok=True)
    with open(args.out, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=1)
    n_k = sum(len(e.get("stocks") or []) for e in by_sector.values())
    n_q = sum(len(e.get("kosdaqStocks") or []) for e in by_sector.values())
    print(f"  Updated {args.out} - {len(by_sector)} sectors, "
          f"KOSPI {n_k} / KOSDAQ {n_q} stocks (top {TOP_PER_SECTOR}, 갱신: {sorted(covered)})")


if __name__ == "__main__":
    main()
